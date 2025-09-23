# tabs/results_tab.py

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import numpy as np
import os
import shutil
import cv2
import json

# Matplotlib imports for the interactive plot
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)

# Try to import openpyxl for Excel export
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ResultsTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        
        self.results_data = None
        self.current_pil_image = None
        self.photo_image = None
        self._resize_img_job_id = None
        self._resize_plot_job_id = None
        
        self.brightness_var = tk.DoubleVar(value=0)
        self.contrast_var = tk.DoubleVar(value=1.0)
        
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)
        
        self.main_pane = ttk.PanedWindow(self, orient=tk.VERTICAL)
        self.main_pane.grid(row=0, column=0, sticky="nsew", padx=10, pady=(10, 5))
        
        self.create_top_pane()
        self.create_bottom_pane()
        self.create_export_section()

        self.brightness_slider.config(command=self.apply_brightness_contrast)
        self.contrast_slider.config(command=self.apply_brightness_contrast)

    def create_top_pane(self):
        top_frame = ttk.Frame(self.main_pane)
        self.main_pane.add(top_frame, weight=3)
        top_pane = ttk.PanedWindow(top_frame, orient=tk.HORIZONTAL)
        top_pane.pack(expand=True, fill='both')
        tree_frame = ttk.LabelFrame(top_pane, text="Detected Wells", padding=5)
        top_pane.add(tree_frame, weight=1)
        self.tree = ttk.Treeview(tree_frame, columns=("Well ID", "Intensity", "Frame #"), show="headings")
        self.tree.heading("Well ID", text="Well ID")
        self.tree.heading("Intensity", text="Intensity")
        self.tree.heading("Frame #", text="Frame #")
        self.tree.column("Well ID", width=80, anchor=tk.CENTER)
        self.tree.column("Intensity", width=120, anchor=tk.CENTER)
        self.tree.column("Frame #", width=100, anchor=tk.CENTER)
        self.tree.pack(expand=True, fill='both')
        self.tree.bind('<<TreeviewSelect>>', self.on_well_select)
        preview_frame = ttk.LabelFrame(top_pane, text="Preview", padding=5)
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)
        top_pane.add(preview_frame, weight=3)
        self.preview_label = ttk.Label(preview_frame, text="No results loaded.", anchor=tk.CENTER)
        self.preview_label.grid(row=0, column=0, columnspan=2, sticky="nsew")
        self.preview_label.bind('<Configure>', self.on_preview_resize)
        ttk.Label(preview_frame, text="Brightness:").grid(row=1, column=0, sticky='e', pady=(5,0))
        self.brightness_slider = ttk.Scale(preview_frame, from_=-100, to=100, orient=tk.HORIZONTAL, variable=self.brightness_var, state=tk.DISABLED)
        self.brightness_slider.grid(row=1, column=1, sticky='ew', padx=5, pady=(5,0))
        ttk.Label(preview_frame, text="Contrast:").grid(row=2, column=0, sticky='e')
        self.contrast_slider = ttk.Scale(preview_frame, from_=0.1, to=3.0, orient=tk.HORIZONTAL, variable=self.contrast_var, state=tk.DISABLED)
        self.contrast_slider.grid(row=2, column=1, sticky='ew', padx=5)

    def create_bottom_pane(self):
        bottom_container = ttk.Frame(self.main_pane)
        self.main_pane.add(bottom_container, weight=2)
        bottom_container.columnconfigure(0, weight=1)
        bottom_container.rowconfigure(0, weight=1)
        bottom_pane = ttk.PanedWindow(bottom_container, orient=tk.HORIZONTAL)
        bottom_pane.grid(row=0, column=0, sticky='nsew')
        text_frame = ttk.LabelFrame(bottom_pane, text="Analysis Details", padding=5)
        bottom_pane.add(text_frame, weight=1)
        text_frame.rowconfigure(0, weight=1)
        text_frame.columnconfigure(0, weight=1)
        self.summary_text = tk.Text(text_frame, height=6, wrap=tk.WORD, state=tk.DISABLED, bg="#2b2b2b", fg="white")
        self.summary_text.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.summary_text.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.summary_text.config(yscrollcommand=scrollbar.set)
        plot_frame = ttk.LabelFrame(bottom_pane, text="Intensity Plot", padding=5)
        bottom_pane.add(plot_frame, weight=2)
        plot_frame.rowconfigure(0, weight=1)
        plot_frame.columnconfigure(0, weight=1)
        plot_frame.bind('<Configure>', self.on_plot_resize)
        plt.style.use('dark_background')
        self.fig = Figure(figsize=(8, 4), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_frame)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky='nsew')
        toolbar = NavigationToolbar2Tk(self.canvas, plot_frame, pack_toolbar=False)
        toolbar.update()
        toolbar.grid(row=1, column=0, sticky='ew')
        
    def create_export_section(self):
        button_frame = ttk.LabelFrame(self, text="Export", padding=10)
        button_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
        self.save_frame_btn = ttk.Button(button_frame, text="Save Selected Frame", state=tk.DISABLED, command=self.save_selected_frame)
        self.save_frame_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        self.save_all_frames_btn = ttk.Button(button_frame, text="Save All Peak Frames", state=tk.DISABLED, command=self.save_all_peak_frames)
        self.save_all_frames_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        self.save_map_btn = ttk.Button(button_frame, text="Save Well Map", state=tk.DISABLED, command=self.save_well_map)
        self.save_map_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        self.save_json_btn = ttk.Button(button_frame, text="Save Data (JSON)", state=tk.DISABLED, command=self.save_data_json)
        self.save_json_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        self.save_excel_btn = ttk.Button(button_frame, text="Save Data (Excel)", state=tk.DISABLED, command=self.save_data_excel)
        self.save_excel_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        self.clear_btn = ttk.Button(button_frame, text="Clear Results", state=tk.DISABLED, command=self.clear_results)
        self.clear_btn.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        
    def on_preview_resize(self, event=None):
        if self._resize_img_job_id:
            self.after_cancel(self._resize_img_job_id)
        self._resize_img_job_id = self.after(100, self.apply_brightness_contrast)

    def on_plot_resize(self, event=None):
        if self._resize_plot_job_id:
            self.after_cancel(self._resize_plot_job_id)
        self._resize_plot_job_id = self.after(300, self._update_intensity_plot)
            
    def display_results(self, results_package):
        self.clear_results()
        self.results_data = results_package
        
        # --- MODIFIED: Update tree headings and content based on result type ---
        is_video = self.results_data.get('total_frames', 1) > 1
        self.tree.heading("Intensity", text="Peak Intensity" if is_video else "Intensity")
        for item in self.results_data['numerical_data']:
            self.tree.insert('', tk.END, values=(
                f"Well {item['well_id']+1}",
                f"{item['intensity']:.2f}",
                item['frame'] if is_video else "-"
            ))
            
        self.summary_text.config(state=tk.NORMAL)
        self.summary_text.insert(tk.END, self.results_data['summary_text'])
        self.summary_text.config(state=tk.DISABLED)
        self._update_intensity_plot()
        if self.tree.get_children():
            first_item = self.tree.get_children()[0]
            self.tree.selection_set(first_item)
            self.tree.focus(first_item)
        self.set_controls_state(tk.NORMAL)

    def _update_intensity_plot(self):
        # --- MODIFIED: Show a bar chart for images, line plot for videos ---
        if not self.results_data:
            return
        self.ax.clear()
        
        is_video = self.results_data.get('total_frames', 1) > 1
        
        if is_video:
            # Existing time-series plot for videos
            intensity_data = self.results_data['intensity_data']
            peak_results = self.results_data['numerical_data']
            sample_rate = self.results_data['sample_rate']
            for i, well_data in enumerate(intensity_data):
                num_samples = len(well_data)
                frame_numbers = np.arange(num_samples) * sample_rate
                self.ax.plot(frame_numbers, well_data, label=f'Well {i+1}')
                for peak in peak_results:
                    if peak['well_id'] == i:
                        self.ax.plot(peak['frame'], peak['intensity'], 'ro', markersize=8)
                        break
            self.ax.set_title('Well Intensity vs. Time', fontsize=12)
            self.ax.set_xlabel('Frame Number')
            self.ax.set_ylabel('Brightness')
            self.ax.legend()
        else:
            # New bar chart for single images
            results = self.results_data['numerical_data']
            well_labels = [f"Well {r['well_id'] + 1}" for r in results]
            intensities = [r['intensity'] for r in results]
            self.ax.bar(well_labels, intensities, color='cyan')
            self.ax.set_title('Well Intensity', fontsize=12)
            self.ax.set_xlabel('Well ID')
            self.ax.set_ylabel('Brightness')
            self.ax.tick_params(axis='x', rotation=45)

        self.ax.grid(True, linestyle='--', alpha=0.6)
        self.fig.tight_layout()
        self.canvas.draw()
        
    def on_well_select(self, event):
        selected_items = self.tree.selection()
        if not selected_items:
            return
        item_data = self.tree.item(selected_items[0])
        well_id_str = item_data['values'][0]
        well_index = int(well_id_str.split(' ')[1]) - 1
        if self.results_data:
            pil_image = self._generate_peak_frame_image(well_index)
            if pil_image:
                self._load_image_to_preview(pil_image)

    def _generate_peak_frame_image(self, well_index):
        # --- MODIFIED: Handle both video and single image sources ---
        if not self.results_data:
            return None
        
        source_path = self.results_data['video_path'] # This key holds path for both videos and images
        peak_info = self.results_data['numerical_data'][well_index]
        roi = self.results_data['well_rois'][well_index]
        is_video = self.results_data.get('total_frames', 1) > 1
        frame = None

        if is_video:
            frame_idx = peak_info['frame']
            cap = cv2.VideoCapture(source_path)
            if not cap.isOpened():
                return None
            cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
            ret, frame = cap.read()
            cap.release()
            if not ret:
                return None
        else: # Is single image
            frame = cv2.imread(source_path)
            if frame is None:
                return None

        (x, y, w, h) = roi
        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
        info_text = f"Well {well_index+1}: {peak_info['intensity']:.2f}"
        cv2.putText(frame, info_text, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        return Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))

    def _load_image_to_preview(self, pil_image):
        self.current_pil_image = pil_image
        self.brightness_var.set(0)
        self.contrast_var.set(1.0)
        self.apply_brightness_contrast()
    
    def apply_brightness_contrast(self, event=None):
        if self.current_pil_image is None:
            return
        brightness = self.brightness_var.get()
        contrast = self.contrast_var.get()
        img_np = np.array(self.current_pil_image).astype(np.int16)
        adjusted_np = np.clip(img_np * contrast + brightness, 0, 255).astype(np.uint8)
        adjusted_pil = Image.fromarray(adjusted_np)
        self.preview_label.update_idletasks()
        lw, lh = self.preview_label.winfo_width(), self.preview_label.winfo_height()
        if lw > 1 and lh > 1:
            adjusted_pil.thumbnail((lw, lh), Image.Resampling.LANCZOS)
        self.photo_image = ImageTk.PhotoImage(image=adjusted_pil)
        self.preview_label.config(image=self.photo_image, text="")

    def save_selected_frame(self):
        if not self.current_pil_image:
            return
        filepath = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG", "*.png"), ("JPEG", "*.jpg")])
        if not filepath:
            return
        try:
            brightness = self.brightness_var.get()
            contrast = self.contrast_var.get()
            img_np = np.array(self.current_pil_image).astype(np.int16)
            adjusted_np = np.clip(img_np * contrast + brightness, 0, 255).astype(np.uint8)
            Image.fromarray(adjusted_np).save(filepath)
            messagebox.showinfo("Success", f"Image saved successfully to: {filepath}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save the image: {e}")

    def save_all_peak_frames(self):
        if not self.results_data:
            return
        directory = filedialog.askdirectory(title="Select Directory to Save All Frames")
        if not directory:
            return
        try:
            num_wells = len(self.results_data['numerical_data'])
            for i in range(num_wells):
                pil_image = self._generate_peak_frame_image(i)
                if pil_image:
                    filename = f'peak_frame_well_{i+1}.png'
                    dest_path = os.path.join(directory, filename)
                    pil_image.save(dest_path)
            messagebox.showinfo("Success", f"Successfully saved {num_wells} peak frames to:\n{directory}")
        except Exception as e:
            messagebox.showerror("Save Error", f"An error occurred: {e}")
            
    def save_well_map(self):
        if not self.results_data:
            return
        filepath = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Image", "*.png")])
        if not filepath:
            return
        try:
            max_frame = self.results_data['max_intensity_frame']
            rois = self.results_data['well_rois']
            annotated_image = cv2.cvtColor(max_frame, cv2.COLOR_GRAY2BGR)
            for i, (x, y, w, h) in enumerate(rois):
                cv2.rectangle(annotated_image, (x, y), (x+w, y+h), (0, 255, 0), 2)
                cv2.putText(annotated_image, f'Well {i+1}', (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            Image.fromarray(cv2.cvtColor(annotated_image, cv2.COLOR_BGR2RGB)).save(filepath)
            messagebox.showinfo("Success", f"Well map saved successfully to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save the well map: {e}")

    def _generate_default_filename(self, extension):
        # --- MODIFIED: New helper to create descriptive default filenames ---
        if not self.results_data:
            return f"analysis.{extension}"
        
        base_name = os.path.splitext(os.path.basename(self.results_data.get("video_path", "analysis")))[0]
        metric = self.results_data.get("metric_mode", "intensity")
        
        is_video = self.results_data.get('total_frames', 1) > 1
        if is_video:
            # A more robust solution would pass the calibration value to the results_data package
            return f"{base_name}_{metric}_video.{extension}"
        else:
            return f"{base_name}_{metric}_image.{extension}"

    def save_data_json(self):
        if not self.results_data:
            return
        
        default_name = self._generate_default_filename("json")
        filepath = filedialog.asksaveasfilename(
            initialfile=default_name,
            defaultextension=".json",
            filetypes=[("JSON", "*.json")]
        )
        if not filepath:
            return
            
        try:
            # --- MODIFIED: Conditionally include intensity_timeseries ---
            is_video = self.results_data.get('total_frames', 1) > 1
            
            peak_results_native = [
                {
                    'well_id': int(item['well_id']),
                    'intensity': float(item['intensity']),
                    'frame': int(item['frame']),
                    'metric_mode': item['metric_mode']
                } for item in self.results_data.get("numerical_data", [])
            ]

            output_data = {
                "analysis_info": {
                    "source_filename": os.path.basename(self.results_data.get("video_path", "N/A")),
                    "analysis_timestamp": self.results_data.get("analysis_timestamp"),
                    "total_frames": int(self.results_data.get("total_frames", 0)),
                    "fps": float(self.results_data.get("fps", 0.0)),
                    "duration_seconds": float(self.results_data.get("duration_seconds", 0.0)),
                    "metric_mode_used": self.results_data.get("metric_mode"),
                    "sample_rate": int(self.results_data.get("sample_rate", 1))
                },
                "peak_results": peak_results_native,
                "well_rois": [
                    {"well_id": i, "x": r[0], "y": r[1], "width": r[2], "height": r[3]}
                    for i, r in enumerate(self.results_data.get("well_rois", []))
                ]
            }
            
            if is_video:
                output_data["intensity_timeseries"] = {
                    f"well_{i+1}": [float(val) for val in data]
                    for i, data in enumerate(self.results_data.get("intensity_data", []))
                }

            with open(filepath, 'w') as f:
                json.dump(output_data, f, indent=4)
                
            messagebox.showinfo("Success", f"JSON data saved successfully to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save the JSON file: {e}")
            
    def save_data_excel(self):
        if not self.results_data:
            return

        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "Dependency Missing",
                "The 'openpyxl' library is required to export to Excel.\n"
                "Please install it using: pip install openpyxl"
            )
            return

        default_name = self._generate_default_filename("xlsx")
        filepath = filedialog.asksaveasfilename(
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")]
        )
        if not filepath:
            return
            
        try:
            wb = openpyxl.Workbook()
            ws_info = wb.active
            ws_info.title = "General Info"
            ws_avg = wb.create_sheet("Average Intensity Data")
            ws_peak = wb.create_sheet("Peak Intensity Data")

            info = {
                "Source Filename": os.path.basename(self.results_data.get("video_path", "N/A")),
                "Analysis Timestamp": self.results_data.get("analysis_timestamp"),
                "Total Frames": self.results_data.get("total_frames"),
                "FPS": self.results_data.get("fps"),
                "Duration (seconds)": self.results_data.get("duration_seconds"),
                "Metric Mode for UI": self.results_data.get("metric_mode"),
                "Sample Rate": self.results_data.get("sample_rate")
            }
            ws_info.append(["Parameter", "Value"])
            for key, value in info.items():
                ws_info.append([key, value])

            def _populate_data_sheet(worksheet, data_list, sample_rate):
                if not data_list: return
                header = ["Frame Number"] + [f"Well {i+1}" for i in range(len(data_list))]
                worksheet.append(header)
                max_len = max(len(col) for col in data_list) if data_list else 0
                for i in range(max_len):
                    frame_num = i * sample_rate
                    row_data = [data_list[j][i] if i < len(data_list[j]) else None for j in range(len(data_list))]
                    worksheet.append([frame_num] + row_data)

            sample_rate = self.results_data.get("sample_rate", 1)
            _populate_data_sheet(ws_avg, self.results_data.get("average_intensity_data"), sample_rate)
            _populate_data_sheet(ws_peak, self.results_data.get("peak_intensity_data"), sample_rate)

            wb.save(filepath)
            messagebox.showinfo("Success", f"Excel data saved successfully to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save the Excel file: {e}")

    def clear_results(self):
        self.results_data = None
        self.current_pil_image = None
        self.photo_image = None
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.preview_label.config(image='', text="No results loaded.")
        self.summary_text.config(state=tk.NORMAL)
        self.summary_text.delete('1.0', tk.END)
        self.summary_text.config(state=tk.DISABLED)
        self.brightness_var.set(0)
        self.contrast_var.set(1.0)
        self.ax.clear()
        self.ax.set_title('Intensity Plot')
        self.ax.set_xlabel('')
        self.ax.set_ylabel('')
        self.canvas.draw()
        self.set_controls_state(tk.DISABLED)

    def set_controls_state(self, state):
        # --- MODIFIED: Conditionally disable frame-saving buttons ---
        is_video = False
        if state == tk.NORMAL and self.results_data:
            is_video = self.results_data.get('total_frames', 1) > 1

        self.brightness_slider.config(state=state)
        self.contrast_slider.config(state=state)
        self.save_map_btn.config(state=state)
        self.save_json_btn.config(state=state)
        self.save_excel_btn.config(state=state)
        self.clear_btn.config(state=state)
        
        # Only enable frame-saving buttons if it was a video analysis
        self.save_frame_btn.config(state=tk.NORMAL if is_video else tk.DISABLED)
        self.save_all_frames_btn.config(state=tk.NORMAL if is_video else tk.DISABLED)

    def cleanup(self):
        pass